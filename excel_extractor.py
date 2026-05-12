import openpyxl
import io

def extract_text_from_excel(file_bytes, filename):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        text = "EXCEL FILE: " + filename + "\n"
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += "\n[Sheet: " + sheet_name + "]\n"
            for row in ws.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    text += " | ".join(row_values) + "\n"
        return text
    except Exception as e:
        return "Excel extraction failed: " + str(e)
